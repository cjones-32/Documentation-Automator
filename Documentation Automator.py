# Important to PIP install openpyxl to be able to import it
import os
import sys
import shutil
import re
import csv
import openpyxl
import zipfile
import time
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment

import shutil # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED

###################################################################################################
###################################################################################################
###############################     Unzip the packaged project     ################################
###################################################################################################
###################################################################################################

# Global Variables
zip_pcb_number = ''
files_found = []

# Delete full folder or individual files
###################################################################################################
def remove_file_dir(path):
    if os.path.exists(path):
        if os.path.isdir(path):
            shutil.rmtree(path)
        elif os.path.isfile(path):
            os.remove(path)


# Walk through all the subfolders of the folder passed to it and record them to be moved
###################################################################################################
# Directory is the current folder full path that is being walked through, ie the extracted zip top level or extracted zip project outputs
# prj_folder is the folder location and name for the full project file that is final name ready for unreleased/released

def walk_folder(directory, prj_folder):

    global files_found

    # known folders in the project directory
    project_folders = ['build request',
                       'design notes & rules',
                       'design notes and rules',
                       'mechanicals',
                       'reports',
                       'settings',
                       'source'
                       ]
    
    high_level_folders = ['cam',
                        'email',
                        'mfg-data',
                        'planr testpoints',
                        'testpoints for todor'
                        ]

    known_junk = ['^project logs for (.*)$',
                '.gitignore$',
                '.git$',
                '^' + zip_pcb_number + '.prjpcbstructure$',
                '^status report.txt$',
                '^history$',
                '^design rule check(.*)$',
                '^(.*).was$'
                ]

    project_files = ['^' + zip_pcb_number + '.prjpcb$',
                     '^' + zip_pcb_number + '.annotation$',
                     '^' + zip_pcb_number + '.PrjPcbVariants$',
                     '^' + zip_pcb_number[0:5] + '5.' + zip_pcb_number[7:9] + '....assembly drawing.pdf$',
                     '^' + zip_pcb_number[0:5] + '5.' + zip_pcb_number[7:9] + '....schematic.pdf$',
                     '^' + zip_pcb_number[0:5] + '5.' + zip_pcb_number[7:9] + '....assembly bom.pdf$',
                     '^' + zip_pcb_number[0:5] + '5.' + zip_pcb_number[7:9] + '....pdf$'
                     ]

    # Records if the file is already in the list
    previously_found = False
    
    # Work through all project outputs folder for project folder level
    for folder in os.listdir(directory):
        # If the top level file is junk, delete it checking against regular expression
        skip = [folder for junk in known_junk if re.search(junk, folder, re.IGNORECASE)]
        # Move is junk just delete it.
        if folder in skip:
            remove_file_dir(f'{directory}\\{folder}')
        # Check if its high level or low level
        if folder.lower() in project_folders or folder.lower() in high_level_folders:
            for file in os.listdir(f'{directory}\\{folder}'):
                # Same as above, but check each folder
                skip = [file for junk in known_junk if re.search(junk, file, re.IGNORECASE)]
                # Move is junk just delete it.
                if file in skip:
                    remove_file_dir(f'{directory}\\{folder}\\{file}')
                # Otherwise process as normal
                else:
                    # Set destination for the file
                    if folder.lower() in project_folders:
                        destination = f'{prj_folder}\\{zip_pcb_number}_Prototype\\{folder}'
                    elif folder.lower() in high_level_folders:
                        destination = f'{prj_folder}\\{folder}'
                    else:
                        print('Destination not defined')
                        destination = 'Not Defined'
                    previously_found = False
                    # Walk through all files found
                    for item in files_found:
                        # Check if the file is already in the list to move
                        if item['file'] == file and item['destination'] == destination:
                            previously_found = True
                            # Check if the one stored already is older
                            if item['file time'] < os.path.getmtime(f'{directory}\\{folder}\\{file}'):
                                # If it is, delete the old one and make newer one to move
                                remove_file_dir(item['directory'] + f'\\{folder}\\{file}')
                                # If the folder is empty after deleting the old file, delete the folder
                                if not os.listdir(item['directory'] + f'\\{folder}'):
                                    remove_file_dir(item['directory'] + f'\\{folder}')
                                item['directory'] = directory
                            # It's older, just delete it
                            else:
                                remove_file_dir(f'{directory}\\{folder}\\{file}')
                    # File not previously found, add it
                    if previously_found is False:
                        files_found.append({'file'       : file, # File to be moved
                                            'directory'  : directory, # Directory of where it's moving from ie extracted zip
                                            'folder'     : folder, # Folder it belongs into ex. reports
                                            'destination': destination, # Directory of where its going to ie project folder
                                            'file time'  : os.path.getmtime(f'{directory}\\{folder}\\{file}'), # Date of the file to be moved
                                            'outdated'      : 'Unknown'}) # Will be used to record if the zip file is newer then what may already exist
        # Handle the project files that are not in a folder.
        else:
            previously_found = False
            for file in project_files:
                if re.search(file, folder, re.IGNORECASE):
                    for item in files_found:
                        # Check if the file is already in the list to move
                        if item['file'] == folder and item['destination'] == prj_folder:
                            previously_found = True
                            # Check if the one stored already is older
                            if item['file time'] < os.path.getmtime(f'{directory}\\{folder}'):
                                # If it is, delete the old one and make newer one to move
                                remove_file_dir(item['directory'] + f'\\{folder}')
                                # If the folder is empty after deleting the old file, delete the folder
                                if not os.listdir(item['directory'] + f'\\{folder}'):
                                    remove_file_dir(item['directory'] + f'\\{folder}')
                                item['directory'] = directory
                            # It's older, just delete it
                            else:
                                remove_file_dir(f'{directory}\\{folder}')
                    if previously_found is False:
                        files_found.append({'file'       : folder, # File to be moved NOTE DIFFERENT
                                            'directory'  : directory, # Directory of where it's moving from ie extracted zip NOTE DIFFERENT
                                            'folder'     : '', # Folder it belongs into ex. reports
                                            'destination': f'{prj_folder}\\{zip_pcb_number}_Prototype', # Directory of where its going to ie project folder NTOE DIFFERENT
                                            'file time'  : os.path.getmtime(f'{directory}\\{folder}'), # Date of the file to be moved
                                            'outdated'   : 'Unknown'}) # Will be used to record if the zip file is newer then what may already exist

# Move all files as needed in the found files from the walk function
###################################################################################################
def move_files():
    
    global zip_pcb_number
    global files_found


    # Walk through all the files to be moved
    for file in files_found:

        if file['folder']:
            source = file['directory'] + '\\' + file['folder'] + '\\' + file['file']
        else:
            source = file['directory'] + '\\' + file['file']

        destination = file['destination'] + '\\' + file['file']

        # Making current folder if its not there
        if not os.path.exists(file['destination']):
            os.mkdir(file['destination'])
        
        # See if it already is there
        if os.path.exists(file['destination'] + '\\' + file['file']):
            # It's there, check if zip is newer then current file
            if os.path.getmtime(source) > os.path.getmtime(destination):
                # It is newer, move it.
                shutil.move(source, destination)
                file['outdated'] = 'False'
            # Project file is newer, record that to prompt user later
            elif os.path.getmtime(source) < os.path.getmtime(destination):
                file['outdated'] = 'True'
            # They are the same, just delete instead of moving or prompting
            else:
                remove_file_dir(source)
                
        # It's not there, move it
        else:
            shutil.move(source, destination)
            file['outdated'] = 'False'
            
        # Try to delete folder
        if not os.listdir(file['directory'] + '\\' + file['folder']):
            os.rmdir(file['directory'] + '\\' + file['folder'])

    # Collect all the files that are out of date
    outdated_files = [file for file in files_found if file['outdated'] == 'True']

    # Only do this if there are outdated files
    if outdated_files:

        print()
        # Show all the out of date files
        for file in outdated_files:
            print(file['destination'] + '\\' + file['file'])

        # Overwrite if allowed
        overwrite = 'Unknown'
        while overwrite not in ('y', 'yes', 'n', 'no'):
            overwrite = input('\nALERT! Existing project folder has newer files then the zip! Would you like to overwrite with the old zip file? ').lower()
            # Move the files with a yes prompt
            if response_unpackage in ('y', 'yes'):
                for file in outdated_files:
                    shutil.move(file['directory'] + '\\' + file['folder'] + '\\' + file['file'], file['destination'] + '\\' + file['file'])
            # Skip the files
            else:
                print('Files skipped')

# Unzip and create the project folders as needed
###################################################################################################
def unpack_project(zip_path):
    # Needed variables
    global zip_pcb_number
    global files_found

    # Extracted folder where files are pulled from then deleted
    extracted_location = ''
    # Project selected by user in extracted location that is used to name the file
    extracted_project = ''
    # Project folder where the final project lives
    project_folder = ''
    # .PrjPCB file selected by user if multiple present
    selected_project = ''
    
    projects_found = []

    # Get the project Directory. Take the Altium project path, rsplit the last \ and take the first string of the 2 new ones
    project_dir = zip_path.rsplit('\\', 1)[0]
    # Store the original working directory to restore later
    owd = os.getcwd()
    # Change the current working directory to that of the project. Make the \ a \\ to allow python to take the slashes literally.
    os.chdir(project_dir.replace('\\', '\\\\'))

    # Take the path, check if whats after the last \ starts with a PCB number
    if re.search('^\d\d\d\dB46\d\d[A-Z]', zip_path.rsplit('\\', 1)[1]):
        # Grab the first 11 characters of the file
        zip_pcb_number = zip_path.rsplit('\\', 1)[1][:10]
    else:
        # Grab the first 11 characters of the file
        zip_pcb_number = zip_path.rsplit('\\', 1)[1][:10]
        # Alert user PCB number could not be found
        print(f'PCB number not found at start of zip, folders may be name incorrect.\nPCB number used is - {zip_pcb_number}')

    # Get date and time of selected zip
    zip_date = os.path.getmtime(zip_path)
    # Used to find the newest file
    temp_time = os.path.getmtime(zip_path)
    temp_file = zip_path
    # Get each zip file in the directory that starts with the same PCB Number
    for file in os.listdir('.\\'):
        if re.search('^' + zip_pcb_number + '.*\.zip$', file, re.IGNORECASE):
            # If it is newer, ask if that should be unpackaged instead.
            if os.path.getmtime(file) > temp_time:
                temp_file = file
                temp_time = os.path.getmtime(file)

    if zip_path is not temp_file:
        response_replace = ''
        while response_replace not in ('y', 'yes', 'n', 'no'):
            response_replace = input(f'\n{temp_file} - Newer project zip found. Unpackage this instead? ').lower()
        # If yes, replace the zip path the user submitted with the new one.
        if response_replace in ('y', 'yes'):
            zip_path = os.getcwd() + '\\' + temp_file

    # See if the extracted file already exists
    if os.path.exists(zip_path.rsplit('.', 1)[0]):
        folder_inc = 0
        # If it does count up till you can name the folder with a number
        while os.path.exists(zip_path.rsplit('.', 1)[0]  + ' (' + str(folder_inc) + ')'):
            folder_inc += 1
        extracted_location = zip_path.rsplit('.', 1)[0]  + ' (' + str(folder_inc) + ')'
            
    else:
        # Extracted location will be same name
        extracted_location = zip_path.rsplit('.', 1)[0]

    # Unzip the project
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extracted_location)
        
    # Make the modified date same as what it was in the zip
    # https://stackoverflow.com/questions/9813243/extract-files-from-zip-file-and-retain-mod-date
    for f in zipfile.ZipFile(zip_path, 'r').infolist():
        # path to this extracted f-item
        fullpath = os.path.join(extracted_location, f.filename)
        # still need to adjust the dt o/w item will have the current dt
        date_time = time.mktime(f.date_time + (0, 0, -1))
        # update dt
        os.utime(fullpath, (date_time, date_time))

    # Go through all files in the top directory, if its a project file add it to projects found
    projects_found = [file for file in os.listdir(extracted_location) if re.search('\.PrjPcb$', file, re.IGNORECASE)]
    # If no project files found, alert user and exit
    if projects_found is None:
        input('\nNo Altium project found in top level of zip, Press enter to exit ')
        exit()
    # If more then one project file is found, print them all and alert user
    elif len(projects_found) > 1:
        # Record the index number of the project they want to use
        project_choice = -1
        # While the input is not valid, keep asking
        while 0 >= int(project_choice) or int(project_choice) > len(projects_found):
            print('\nMultiple project files found, please select from below:')
            # Print the projects with an index number
            for idx, project in enumerate(projects_found):
                print("{}) {}".format(idx+1,project))
            project_choice = input('Enter the number of the project you would like to use: ')
            if project_choice.isdigit():
                project_choice = int(project_choice)
            else:
                project_choice = -1
        selected_project = projects_found[project_choice-1]
        extracted_project = f'{extracted_location}\\{selected_project}'

    else:
        # Record the full location and file name of the only project
        extracted_project = f'{extracted_location}\\{projects_found[0]}'
        selected_project = projects_found[0]
        
    # Get the folder name from the project file
    with open(extracted_project, 'r', encoding = 'UTF-8') as project_file:
        lines = project_file.readlines()
        for i in range(0, len(lines)):
            line = lines[i]
            # If the folder description line is found, verify the next is value and store it
            if line == 'Name=A104_Project_Info_Folder_Description\n' or line == 'Name=A104_Project_Info_PCB_Folder_Description\n':
                if lines[i+1].split('=')[0] == 'Value':
                    project_folder = zip_pcb_number + '_' + lines[i+1].split('=')[1].rstrip()
        # If the folder description line was not found, use SAP Description
        if project_folder == '':
            print('Folder name parameter not found, using SAP description instead.')
            for i in range(0, len(lines)):
                line = lines[i]
                # If the folder description line is found, verify the next is value and store it
                if line == 'Name=A103_Project_Info_PCB_SAP_Description\n':
                    if lines[i+1].split('=')[0] == 'Value':
                        project_folder = zip_pcb_number + '_' + lines[i+1].split('=')[1].rstrip()

    #If the user didnt select a project when multipe were found, alert and exit.
    if extracted_project == '':
        input('\nNo Altium project selected, Press enter to exit ')
        exit()

    # See if the correct folder name is there, if not make it and other needed folders.
    try:
        if not os.path.exists(project_folder):
            os.mkdir(project_folder)
            os.mkdir(f'{project_folder}\\_Vaulted')
            os.mkdir(f'{project_folder}\\{zip_pcb_number}_Prototype')
        if not os.path.exists(f'{project_folder}\\Deleted'):
            os.mkdir(f'{project_folder}\\Deleted')
        if not os.path.exists(f'{project_folder}\\Deleted\\From Unpackage'):
            os.mkdir(f'{project_folder}\\Deleted\\From Unpackage')
    # Likely error caused by bad character in folder name parameter
    except OSError as e:
        print ("Error: %s - %s.\n Possible invalid folder character in parameter" % (e.filename, e.strerror))

# Work through the temp files
###################################################################################################

    # Verify project outputs folder is correct
    if os.path.exists(f'{extracted_location}\\Project Outputs for {zip_pcb_number}'):
        outputs_folder = f'{extracted_location}\\Project Outputs for {zip_pcb_number}'
    else:
        outputs_folders = []
        for file in os.listdir(extracted_location):
            if re.search('^Project Outputs for ', file, re.IGNORECASE):
                outputs_folders.append(file)
        if not outputs_folders:
            print('No outputs folder found!! Managed outjob not ran')
        elif len(outputs_folders) == 1:
            print('Project outputs folder incorrectly named. Verify ' + outputs_folders[0] + ' is correct to use')
            outputs_folder = f'{extracted_location}\\{outputs_folders[0]}'
        else:
            selected_output = -1
            while 0 >= int(selected_output) or int(selected_output) > len(outputs_folders):
                print('\nMultiple project output folders found, please select from below:')
                for idx, folder in enumerate(outputs_folders):
                    print("{}) {}".format(idx+1,folder))
                selected_output = input('Enter the number of the projects outputs folder you would like to use: ')
                if selected_output.isdigit():
                    selected_output = int(selected_output)
                else:
                    selected_output = -1
            outputs_folder = f'{extracted_location}\\{outputs_folders[selected_output-1]}'
                
    # Walk through the zip folder files
    walk_folder(extracted_location, project_folder)
    
    # Walk through the zip folder "Project Outputs for..." folder
    walk_folder(outputs_folder, project_folder)

    # Move all the files that were found in the walk
    move_files()

    # Try to delete project outputs folder if it's empty
    try:
        os.rmdir(outputs_folder)
    except:
        False
    
# Cleanup and closing
###################################################################################################

# Try to delete project outputs folder if it's empty
    try:
        os.rmdir(outputs_folder)
    except:
        False

    try:
        os.rmdir(f'{extracted_location}')
    except:
        print('Unable to process all files, temporary zip has been kept with unmanaged files')
        
    selected_project = os.getcwd() + '\\' + project_folder + '\\' + zip_pcb_number + '_Prototype\\' + selected_project
    
    # Change working directory back to default to prevent program from preventing deleting project file
    os.chdir(owd)
    
    # Return the location of the project selected
    return(selected_project.replace('\\\\', '\\'))

###############################################################################################################################

###############################################################################################################################
###############################################################################################################################
##############     Get the project directory, project number, and all assemblies     ##########################################
###############################################################################################################################
###############################################################################################################################

###############################################################################################################################

# Clear the screen and print version.
os.system("cls")
print('Documentation Automator v1.0.0\n')
# The number of assemblies found
assembly_count = 0
# All the assemblies in the project
assemblies = []

# Get the path of the Altium project from user
project_path = input('Enter the full path of your Altium project:\n')

# Strip quotes if present from copying path in windows
if project_path[0] == '"' and project_path[-1] == '"':
    project_path = project_path.strip('"')

# Verify the file is at the location of the user input
if not os.path.isfile(project_path):
    input('\nProject not found. Press enter to exit.')
    exit()
    
# Verify that it is an Altium project file given ie ending in .PrjPcb
if not re.search('(\.PrjPcb)?(\.zip)?$', project_path, re.IGNORECASE):
    input('\nFile linked to is not an Altium project or zip file. Press enter to exit.')
    exit()

if re.search('\.zip$', project_path, re.IGNORECASE):
    response_unpackage = ''
    while response_unpackage not in ('y', 'yes', 'n', 'no'):
        response_unpackage = input('Zipped project found, would you like to unpackage? ').lower()
        if response_unpackage in ('y', 'yes'):
            project_path = unpack_project(project_path)
            print('\nProject unpackaging complete!')
        else:
            input('\nAborted, press enter to exit.')
            exit()
    response_continue = ''
    while response_continue not in ('y', 'yes', 'n', 'no'):
        response_continue = input('\nWould you like to continue documentation cleanup with:\n' + project_path + '? ').lower()
        if response_continue in ('n', 'no'):
            input('Aborted, press enter to exit.')
            exit()
    

# Get the board number.
# Take the Altium project path, rsplit the last \ if present and take the second string of the 2 new ones
if '\\' in project_path:
    pcb_number = project_path.rsplit('\\', 1)[1]

# Then rsplit again by underscore if _v0x is present to come up with board name in the end.
if '_' in pcb_number:
    pcb_number = pcb_number.rsplit('_', 1)[0]
# If _v0x is not present, remove the file extension and check validity.
else:
    pcb_number = pcb_number.rsplit('.', 1)[0]
 
# Verifies that the board number matches the template of 1234B4657A
if not re.search('^\d\d\d\dB46\d\d[A-Z]$', pcb_number):
    input('\nError: Invalid project number, project number found: \'' + pcb_number + '\'. Press enter to exit.')
    exit()
    
# Get the project Directory. Take the Altium project path, rsplit the last \ and take the first string of the 2 new ones
project_dir = project_path.rsplit('\\', 1)[0]
# Store the original working directory to restore later
owd = os.getcwd()
# Change the current working directory to that of the project. Make the \ a \\ to allow python to take the slashes literally.
os.chdir(project_dir.replace('\\', '\\\\'))

# Get assemblies from project file
# Open the project, read its lines, and if it find [ProjectVariantX] verify 2 lines ahead is the description and store its value
with open(project_path, 'r', encoding = 'UTF-8') as project_file:
    lines = project_file.readlines()
    for i in range(0, len(lines)):
        line = lines[i]
        if line == f"[ProjectVariant{assembly_count + 1}]\n":
            if lines[i+2].split('=')[0] == 'Description':
                assembly_count += 1
                assemblies.append(lines[i+2].split('=')[1].rstrip())

print('\nProject Information')
print(  '*******************')
print(str(assembly_count) + ' Assemblie(s) Found In Project File:')
for i in assemblies: print(i)
print()

###################################################################################################
####################     Declare arrays and dictionaries for needed files     #####################
###################################################################################################

# List of file we are going to be manipulating. Will allow ability for user to choose what to keep in future updates
###################################################################################################

# Assembly BOMs found
assembly_boms = []
# SAP BOMs found
sap_boms = []
# Aegis Sync BOMs found
aegis_boms = []

# Lists of files that we want to keep with known names. Allows easy way to adjust what should be kept
###################################################################################################

# List of files for the reports folder that we want to keep
reports_keep = ['^' + pcb_number + '[ ,_]Order[ ,_]Information.xls(x)?$',
                '^' + pcb_number + '[ ,_]Build[ ,_]Request.doc(x)?$',
                '^' + pcb_number + '[ ,_]EE[ ,_]Review.xls(x)?$'
                ]

# List of files for Source folder we want to keep
source_keep = ['^Assy[ ,_]' + pcb_number + '(_v[0-9]+)?.PCBDwf$',
                '^Assy[ ,_]' + pcb_number[0:5] + '5[0-9]' + pcb_number[7:10] + '.PCBDwf$',
                '^PCB[ ,_]' + pcb_number + '(_v[0-9]+)?.PcbDoc$',
                '.SchDoc$'
                ]

# List of Mfg-Data files to keep, adds aegis sync later depending on if excel or text needed.
mfgdata_keep = ['^ODB[ ,_]' + pcb_number + '.zip$'
                 ]

# Left empty, filled in when Gerbers are found. Will allow ability for user to choose what to keep in future updates
gerbers_keep = []

# List of files / folders for CAM folder that are valid
cam_keep = ['^Gerber and Drill$',
              '^' + pcb_number + '_[A-Z][0-9]+.pdf$',
              '^' + pcb_number + '_[A-Z][0-9]+_(RoHS)?(R)?(FLEX)?(PILLAR)?(MCPCB)?(VIPPO)?.zip$',
              '^Spec[ ,_]' + pcb_number + '_[A-Z][0-9]+.dwg$',
             ]

# Dictionary of Altiums Gerber layer file extensions
gerber_ext = {  pcb_number + '.G[0-9]+' : '   -  Mid Layer ',              # .G(int) =  internal layer (int)
              pcb_number + '.GBL' : '  -  Bottom Layer',
              pcb_number + '.GBO' : '  -  Bottom Overlay',
              pcb_number + '.GBP' : '  -  Bottom Paste Mask',
              pcb_number + '.GBS' : '  -  Bottom Solder Mask',
#              pcb_number + '.GD[0-9]+' : '   -  Drill Drawing ',        # .GD(int) = Drill Drawing (int)
#              pcb_number + '.GG[0-9]+' : '   -  Drill Guide ',          # .GG(int) = Drill Guide (int)
#             pcb_number + '.GKO' : '  -  Keep Out Layer',
#              pcb_number + '.GM[0-9]+' : '  -  Mechanical Layer ',      # .GM(int) = Mechanical Layer (int)
               pcb_number + '.GP[0-9]+' : '  -  Internal Plane Layer ', # .GP(int) = Internal Plane Layer (int)
#             pcb_number + '.GPB' : '  -  Pad Master Bottom',
#             pcb_number + '.GPT' : '  -  Pad Master Top',
              pcb_number + '.GTL' : '  -  Top Layer',
              pcb_number + '.GTO' : '  -  Top Overlay',
              pcb_number + '.GTP' : '  -  Top Paste Mask',
              pcb_number + '.GTS' : '  -  Top Solder Mask',
#              pcb_number + '.P[0-9]+' : '  -  Gerber Panels ',        # .P0(int) = Ger Panel (int)
              pcb_number + '.DRR' : '  -  NC Drill Report',
              pcb_number + '.TXT' : '  -  Drill File',
              pcb_number + '-SlotHoles.TXT' : '  -  Slot Drill File',
              pcb_number + '-RoundHoles.TXT' : '  -  Hole Drill File'
              }

###################################################################################################
######################     Find all the relavent project info and files     #######################
###################################################################################################

# Find the assembly BOMs
###################################################################################################

print('Assembly BOMs:')
# For each assembly, see if the its BOM is in the reports folder
for assembly in assemblies:
    found = False
    print(assembly, end = '')
    for file in os.listdir('.\\Reports'):
        if re.search(assembly + '[ ,_]Assembly[ ,_]BOM.xls(x)?', file, re.IGNORECASE):
            found = True
            assembly_boms.append(file)
            reports_keep.append(file)
            print(' Assembly BOM Found')
    if found == False:
        print(' MISSING')
print()

# Find the SAP BOMs
###################################################################################################

print('SAP BOMs:')
# For each assembly, see if the its BOM is in the reports folder
for assembly in assemblies:
    found = False
    print(assembly, end = '')
    for file in os.listdir('.\\Reports'):
        if re.search(assembly + '[ ,_]SAP[ ,_]Import[ ,_]File.xls(x)?', file, re.IGNORECASE):
            found = True
            sap_boms.append(file)
            reports_keep.append(file)
            print(' SAP Import File Found')
    if found == False:
        print(' MISSING')
print()

# Find Gerber layer data from Gerber and Drill folder
###################################################################################################

print('Gerber Files Found:')
# For all files/folders in the Gerber and Drill folder, if it is a file continue
for file in [file for file in os.listdir('..\\Cam\\Gerber and Drill\\') if os.path.isfile('..\\Cam\\Gerber and Drill\\' + file)]:
    # For the key and values in the Gerber dictionary
    for gerber, label in gerber_ext.items():
        # If that file matches the key, keep it and print the label (value)
        if re.search(gerber, file, re.IGNORECASE):
            gerbers_keep.append(file)
            print(file + label, end = '')
            # If its a numbered gerber, print what number to the user after its description
            if file[-1].isdigit():
                # Strip everything except the number from the key, then strip that from the file name
                # find digits at the end, convert to int and group returns the substring that matches
                print(int(re.search(r'\d+$', file).group()))
            else:
                # Prints end line that was withheld until number check
                print()
print()

# Find the Aegis Sync BOMs
###################################################################################################

print('Aegis Sync Excel BOMs:')
# For each assembly, see if the its BOM is in the reports folder
for assembly in assemblies:
    excel = ''
    text = ''
    # Find if the excel and text files exist, if so remember them for later
    for file in os.listdir('..\\Mfg-Data\\'):
        if re.search('Aegis[ ,_]Sync[ ,_]' + assembly + '.xls(x)?', file, re.IGNORECASE):
            excel = file
        elif re.search('Aegis[ ,_]Sync[ ,_]' + assembly + '.txt', file, re.IGNORECASE):
            text = file
            
    # If both txt and excel files are there, see which one is newer
    if excel and text:
        # If the excel is newer, alert user and save only excel to be saved
        if os.path.getmtime('..\\Mfg-Data\\' + excel) >= os.path.getmtime('..\\Mfg-Data\\' + text):
            print(text + ' File Found - OUT OF DATE')
            print(excel + ' File Found')
            aegis_boms.append(excel)
            mfgdata_keep.append(excel)
        # If the text is newer, alert user and save only excel to be saved
        else:
            print(text + ' File Found')
            print(excel + ' File Found - Marked For Deletion')
            mfgdata_keep.append(text)
    # If just the excel file is there add it to be saved
    elif excel:
        print(excel + ' File Found')
        aegis_boms.append(excel)
        mfgdata_keep.append(excel)
    # Just the text BOM was found
    elif text:
        print(text + ' File Found - Excel missing')
        mfgdata_keep.append(text)
    else:
        print(assembly + ' MISSING')

###################################################################################################
#################################     Sort the assembly BOMs     ##################################
###################################################################################################

print('\n\nSorting Assembly BOMs')
print('*********************')

# Ask the user if they would like to sort all assembly BOM
response_all = ''
if len(assembly_boms) > 1:
    while response_all not in ('y', 'yes', 'n', 'no'):
        response_all = input('Would you like to sort all assembly BOMs?: ').lower()
        print()
else:
    response_all = 'no'

# Do it for all assembly BOMs found
for assembly_bom in assembly_boms:

    # Ask the user if they would like to sort each assembly BOM if they didnt want to do all
    response2 = ''
    if re.search('^n(o)?$', response_all, re.IGNORECASE):
            while response2 not in ('y', 'yes', 'n', 'no'):
                response2 = input(f'Would you like to sort {assembly_bom}?: ').lower()
    if re.search('^n(o)?$', response2, re.IGNORECASE) and re.search('^n(o)?$', response_all, re.IGNORECASE):
        print(assembly_bom + ' Skipped\n')
        continue
    
    print('Sorting ' + assembly_bom + '...')
    # Open the excel file and go in the first sheet
    wb = openpyxl.load_workbook('.\\Reports\\' + assembly_bom)
    sheet = wb['Sheet1']

    # Set up fonts
    # Font for component section declaration 
    comp_row_font  = NamedStyle(name = 'comp_row_font')
    comp_row_font.font = Font(name = 'Verdana', bold = True, size = 14)
    # Font for the section headers
    comp_header_font  = NamedStyle(name = 'comp_header_font')
    comp_header_font.font = Font(name = 'Verdana', bold = True, size = 12, underline = 'single')
    # Font for default lines
    default_font  = NamedStyle(name = 'default_font')
    default_font.font = Font(name = 'Verdana', size = 12)
    default_font.alignment = Alignment(horizontal = 'left', vertical = 'top')
    # Something to tell when the BOM starts
    header_row = 0
    # A list that will contain dictionaries of the parts
    bom_content = []
    # Tells how far the BOM needs to be moved up to overwrite the template headers
    bom_offset = 14

# Read the cells
###################################################################

    # Scan through each row in the BOM excel from 1 to end
    for row in range(1, sheet.max_row + 1):
        # Column 1 is part number, 2 is quantity etc. Store the values in these variables
        part_number = sheet.cell(row, 1).value
        quantity    = sheet.cell(row, 2).value
        description = sheet.cell(row, 3).value
        designator  = sheet.cell(row, 4).value
        layer       = sheet.cell(row, 5).value
        fitted      = sheet.cell(row, 6).value
        # If its previously found the header, and the part number is not None, add the part
        if header_row != 0 and part_number is not None:
            # Handle layer None components, ask for input until valid response
            if layer == 'None':
                while 1:
                    response = input(f'{part_number} | {description} | is layer None, does it belong on top or bottom side?: ')
                    if re.search('^t(op)?( )?(side)?$', response, re.IGNORECASE):
                        layer = 'Top'
                        break
                    if re.search('^b(ottom)?( )?(side)?$', response, re.IGNORECASE):
                         layer = 'Bottom'
                         break
            # Add the part as a dictionary to the list of parts. Make them strings to allow sorting to work.
            bom_content.append({'part_number': str(part_number),
                                'quantity'   : quantity,
                                'description': str(description),
                                'designator' : str(designator),
                                'layer'      : str(layer),
                                'fitted'     : str(fitted)})
        # If the row equals the header Altium uses, singal that the header has been found
        if str(part_number) + str(quantity) + str(description) + str(designator) + str(layer) + str(fitted) == 'LibRefQuantityDescriptionDesignatorLayerFitted':
            header_row = row
    # If the header was not found, it was probably already messed with. Alert the user
    if header_row == 0:
        print(f'No Altium header row found for {assembly_bom}. Unable to sort...\n')
        continue
    
    # Sort the BOMs, Least priority to most. Sorting doesnt change order of those who match sort key
    bom_content.sort(key = lambda k: k['part_number'])
    bom_content.sort(key = lambda k: k['fitted'])
    bom_content.sort(key = lambda k: k['layer'], reverse = True)

    # Delete the old rows of data 
    sheet.delete_rows(header_row - bom_offset, sheet.max_row)
    
# Write the data to the cells
###################################################################

    current_section = 'None'
    pattern_fill = False
    # Start printing 14 rows up from where the header row was found
    row = header_row - bom_offset
    for part in bom_content:
        # Check if you are in a new section of the BOM, if so print the header
        if current_section != part['layer'] + part['fitted']:
            row += 2
            current_section = part['layer'] + part['fitted']
            # Determine what section of BOM is starting
            if part['layer'] == 'Top': header_string = 'TOP SIDE COMPONENTS'
            elif part['layer'] == 'Bottom': header_string = 'BOTTOM SIDE COMPONENTS'
            else: header_string = 'None Side Components'
            if part['fitted'] == 'Not Fitted': header_string += ' (not installed)'
            # Print the component section row and format it
            sheet.cell(row = row, column = 1).value = header_string
            sheet['A' + str(row)].style = comp_row_font
            # Print the component headers and format it
            sheet.cell(row = row + 1, column = 1).value = 'PART#'
            sheet['A' + str(row + 1)].style = comp_header_font
            sheet.cell(row = row + 1, column = 2).value = 'QTY'
            sheet['B' + str(row + 1)].style = comp_header_font
            sheet.cell(row = row + 1, column = 3).value = 'DESCRIPTION'
            sheet['C' + str(row + 1)].style = comp_header_font
            sheet.cell(row = row + 1, column = 4).value = 'SYMBOL'
            sheet['D' + str(row + 1)].style = comp_header_font
            row += 2
            pattern_fill = False
        # Print the row data for the component, give them default 
        sheet.cell(row = row, column = 1).value = part['part_number']
        sheet['A' + str(row)].style = default_font
        if pattern_fill == True: sheet['A' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 2).value = str(part['quantity'])
        sheet['B' + str(row)].style = default_font
        if pattern_fill == True: sheet['B' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 3).value = part['description']
        sheet['C' + str(row)].style = default_font
        if pattern_fill == True: sheet['C' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 4).value = part['designator']  
        sheet['D' + str(row)].style = default_font
        if pattern_fill == True: sheet['D' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 5).value = part['layer']
        sheet['E' + str(row)].style = default_font
        sheet.cell(row = row, column = 6).value = part['fitted']
        sheet['F' + str(row)].style = default_font
        row += 1
        # Toggle the fill pattern for every other fill
        pattern_fill = not pattern_fill

    # Save the excel
    wb.save('.\\Reports\\' + assembly_bom)
    print(assembly_bom + ' sorting complete\n')
    os.startfile('.\\Reports\\' + assembly_bom)

###################################################################################################
#################################     Clean SAP Import Files     ##################################
###################################################################################################

print('Cleaning SAP Import Files')
print('*************************')

# Ask the user if they would like to sort all assembly BOM
response = ''
if len(sap_boms) > 1:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('Would you like to clean all SAP Import files?: ').lower()
        print()
else:
    response = 'no'

# Do it for all assembly BOMs found
for sap_bom in sap_boms:

    # Flag for if a -PROTO part has been found and removed from SAP BOM.
    proto_found = False

    # List to store the dictionary of data
    sap_content = []

    # Ask the user if they would like to sort each assembly BOM if they didnt want to do all
    response2 = ''
    if re.search('^n(o)?$', response, re.IGNORECASE):
            while response2 not in ('y', 'yes', 'n', 'no'):
                response2 = input(f'Would you like to clean {sap_bom}?: ').lower()
    if re.search('^n(o)?$', response2, re.IGNORECASE) and re.search('^n(o)?$', response, re.IGNORECASE):
        print(sap_bom + ' Skipped\n')
        continue
    
    print('Cleaning ' + sap_bom + '...')
    
    # Open the excel file and go in the first sheet
    wb = openpyxl.load_workbook('.\\Reports\\' + sap_bom)
    sheet = wb['Sheet1']
    
    # Read the data
    # Verify the header row is present
    if sheet.cell(1, 2).value != 'LibRef' or sheet.cell(1, 4).value != 'Quantity':
        print(f'No Altium header row found for {sap_bom}. Unable to sort...\n')
        continue
    
    # For each row of data
    for row in range(2, sheet.max_row + 1):
        # Read the data for the rows and store temporarily
        part_number = sheet.cell(row, 2).value
        # If it's a proto part, skip it
        if re.search('-PROTO$', part_number, re.IGNORECASE):
            # Set flag for if a -PROTO part has been found.
            proto_found = True
            continue
        # If there is a software version found, prompt for version
        if re.search('^\d\d\d\dS\d\d\d\d-(X|S)$', part_number, re.IGNORECASE):
            print('Software found')
            # Get the version from user and confirm
            while 1:
                version = input('Undefined software version found: \'' + part_number.split('-', 1)[0] + '-X\', what version should it be? \n')
                part_number = part_number.split('-', 1)[0] + '-' + version
                verify = input(part_number + ' - Is this correct?')
                if re.search('^y(es)?$', verify, re.IGNORECASE):
                    break
        quantity = sheet.cell(row, 4).value
        # If the part number is not none, add that row tothe list
        if part_number is not None:
            sap_content.append({'part_number' : str(part_number),
                                'quantity'    : quantity
                                })
    # Sort the BOM
    sap_content.sort(key = lambda k: k['part_number'])

    # Delete the old rows of data 
    sheet.delete_rows(1, sheet.max_row)
    
    # Write the data
    row = 1
    for part in sap_content:
        sheet.cell(row = row, column = 1).value = 'L'
        sheet.cell(row = row, column = 2).value = part['part_number']
        sheet.cell(row = row, column = 4).value = part['quantity']
        row += 1

    # Save the data
    wb.save('.\\Reports\\' + sap_bom)
    # Alert user if -PROTO part was found and removed
    if proto_found == True:
        print(sap_bom + ' has a -PROTO part. It has been removed from SAP BOM!')
    
    print(sap_bom + ' cleaning complete\n')
    
###################################################################################################
#################################     Tab Delimit Aegis Sync     ##################################
###################################################################################################

print('Tab Delimit Aegis Sync Files')
print('****************************') 

# Ask the user if they would like to sort all assembly BOM
response = ''
if len(aegis_boms) > 1:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('Would you like to create all Aegis sync files?: ').lower()
        print()
else:
    response = 'no'

# Do it for all assembly BOMs found
for excel in aegis_boms:

    # Ask the user if they would like to sort each assembly BOM if they didn't want to do all
    response2 = ''
    if re.search('^n(o)?$', response, re.IGNORECASE):
            while response2 not in ('y', 'yes', 'n', 'no'):
                response2 = input(f'Would you like to make text file for {excel}?: ').lower()
    if re.search('^n(o)?$', response2, re.IGNORECASE) and re.search('^n(o)?$', response, re.IGNORECASE):
        print(excel + ' Skipped\n')
        continue

    # If the text file already exists, delete it
    if os.path.isfile('..\\Mfg-Data\\' + excel.rsplit('.', 1)[0] + '.txt'):
        try:
            os.remove('..\\Mfg-Data\\' + excel.rsplit('.', 1)[0] + '.txt')
        except OSError as e: 
            print ("Error: %s - %s." % (e.filename, e.strerror))
            
    # Open the excel file and go in the first sheet
    print('Generating text file for ' + excel)
    wb = openpyxl.load_workbook('..\\Mfg-Data\\' + excel)
    sheet = wb['Sheet1']
    # Save each row to text file, tab delimited
    with open('..\\Mfg-Data\\' + excel.rsplit('.', 1)[0] + '.txt', 'w', newline = '') as aegis_text:
        file = csv.writer(aegis_text, delimiter = '\t')
        for row in sheet.rows:
            file.writerow([cell.value for cell in row])

    # Finished with conversion, no need to save excel
    print(excel + ' converted to text\n\n')

    # If the excel is in mfgdata_keep, replace it with the text
    mfgdata_keep = [excel.rsplit('.', 1)[0] + '.txt' if x == (excel) else x for x in mfgdata_keep]

# Print if none edited
if not aegis_boms:
    print('No files to convert\n\n')    
    
###################################################################################################
###################################     Remove Junk Files     #####################################
###################################################################################################

print('***********************')
print('Deleting Unneeded Files')
print('***********************\n')

# Reports Folder
###################################################################################################

print('Reports Folder')
print('**************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
reports_unneeded = []

# Scan through each file in the reports folder
for report_file in os.listdir('.\\Reports\\'):
    valid = False
    # Scan through each file in valid reports list, if it matches the file found mark it valid, if its not valid add it to delete list
    for unneeded_file in [report_file for file in reports_keep if re.search(file, report_file, re.IGNORECASE)]:
        valid = True
    if valid == False:
        reports_unneeded.append(report_file)
        print(report_file)

# Ask if ok to delete all unneeded file from Reports
response = ''
if reports_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Reports? ').lower()
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in reports_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('.\\Reports\\' + unneeded_file):
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('.\\Reports\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    os.remove('.\\Reports\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('.\\Reports\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    shutil.rmtree('.\\Reports\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Reports folder done\n')

# Source Folder
###################################################################################################

print('Source Folder')
print('*************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
source_unneeded = []

# Scan through each file in the source folder
for source_file in os.listdir('.\\Source\\'):
    valid = False
    # Scan through each file in valid source list, if it matches the file found mark it valid, if its not valid add it to delete list
    for unneeded_file in [source_file for file in source_keep if re.search(file, source_file, re.IGNORECASE)]:
        valid = True
    if valid == False:
        source_unneeded.append(source_file)
        print(source_file)

# Ask if OK to delete all unneeded file from Reports
response = ''
if source_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Source? ').lower()
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in source_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('.\\Source\\' + unneeded_file):
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('.\\Source\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    os.remove('.\\Source\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('.\\Source\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    shutil.rmtree('.\\Source\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Source folder done\n')

# CAM Folder
###################################################################################################

print('CAM Folder')
print('**************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
cam_unneeded = []

# Scan through each file in the Cam folder
for cam_file in os.listdir('..\\Cam\\'):
    valid = False
    # Scan through each file in valid cam list, if it matches the file found mark it valid, if its not valid add it to delete list
    for unneeded_file in [cam_file for file in cam_keep if re.search(file, cam_file, re.IGNORECASE)]:
        valid = True
    if valid == False:
        cam_unneeded.append(cam_file)
        print(cam_file)

# Ask if OK to delete all unneeded file from Reports
response = ''
if cam_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from CAM? ').lower()
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in cam_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('..\\Cam\\' + unneeded_file):
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('..\\Cam\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    os.remove('..\\Cam\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('..\\Cam\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    shutil.rmtree('..\\Cam\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Cam folder done\n')

# Gerber and Drill Folder
###################################################################################################

print('Gerber and Drill Folder')
print('***********************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
gerber_unneeded = []

# Prints all the files in Reports folder that are not in the list to keep
for unneeded_file in [file for file in os.listdir('..\\Cam\\Gerber and Drill') if file not in gerbers_keep]:
    gerber_unneeded.append(unneeded_file)
    print(unneeded_file)

# Ask if OK to delete all unneeded file from Reports
response = ''
if gerber_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Gerber and Drill? ').lower()
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in gerber_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('..\\Cam\\Gerber and Drill\\' + unneeded_file):
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('..\\Cam\\Gerber and Drill\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    os.remove('..\\Cam\\Gerber and Drill\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('..\\Cam\\Gerber and Drill\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    shutil.rmtree('..\\Cam\\Gerber and Drill\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Gerber and Drill folder done\n')

# Manufacturing and Data Folder
###################################################################################################

print('Mfg-Data Folder')
print('***************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
mfg_unneeded = []

# Scan through each file in the Mfg-Data folder
for mfgdata_file in os.listdir('..\\Mfg-Data\\'):
    valid = False
    # Scan through each file in valid source list, if it matches the file found mark it valid, if its not valid add it to delete list
    for unneeded_file in [mfgdata_file for file in mfgdata_keep if re.search(file, mfgdata_file, re.IGNORECASE)]:
        valid = True
    if valid == False:
        mfg_unneeded.append(mfgdata_file)
        print(mfgdata_file)
        

# Ask if ok to delete all unneeded file from Reports
response = ''
if mfg_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Mfg-Data? ').lower()
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in mfg_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('..\\Mfg-Data\\' + unneeded_file):
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('..\\Mfg-Data\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    os.remove('..\\Mfg-Data\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    if not os.path.exists('..\\Deleted\\'): # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                        os.mkdir('..\\Deleted\\') # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
                    shutil.move('..\\Mfg-Data\\' + unneeded_file, '..\\Deleted\\' + unneeded_file) # DELTE WHEN DELETED FILE REPOSITORY IS NO LONGER NEEDED
#                    shutil.rmtree('..\\Mfg-Data\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Mfg-Data folder done\n')

# Change working directory back to default to prevent program from preventing deleting project file
os.chdir(owd)
input('Documentation cleanup complete!! Press enter to exit.')
