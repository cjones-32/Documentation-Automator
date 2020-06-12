# Important to PIP install openpyxl to be able to import it
import os
import sys
import shutil
import re
import csv
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill

###################################################################################################
##############     Get the project directory, project number, and all assemblies     ##############
###################################################################################################

# The number of assemblies found
assembly_count = 0
# All the assemblies in the project
assemblies = []

# Get the path of the Altium project from user
project_path = input('Enter the full path of your Altium project:\n')

# Strip quotes if present from copying path in windows
if project_path[0] == '"' and project_path[-1] == '"':
    project_path = project_path.strip('"')

# Verify that it is an Altium project file given ie ending in .PrjPcb
if not re.search('\.PrjPcb$', project_path, re.IGNORECASE):
    print('File linked to is not an Altium project')
    exit()

# Verify the file is at the location of the user input
if not os.path.isfile(project_path):
    print('Project not found')
    exit()

# Get the board number. Take the Altium project path, rsplit the last \ and take the second string of the 2 new ones
# Then rsplit again by underscore to come up with board name in the end.
i = project_path.rsplit('\\', 1)[1]
pcb_number = i.rsplit('_', 1)[0]

# Verifies that the board number matches the template of 1234B4657A
if not re.search('^\d\d\d\dB46\d\d[A-Z]$', pcb_number):
    print('Error determining the project number, project number found ', + pcb_number)
    exit()
    
# Get the project Directory. Take the Altium project path, rsplit the last \ and take the first string of the 2 new ones
project_dir = project_path.rsplit('\\', 1)[0]
# Store the original working directory to restore later
owd = os.getcwd()
# Change the current working directory to that of the project. Make the \ a \\ to allow python to take the slashes literally.
os.chdir(project_dir.replace('\\', '\\\\'))

# Get assemblies from project file
# Open the project, read its lines, and if it find [ProjectVariantX] verify 2 lines ahead is the description and store its value
with open(project_path, 'r', encoding = 'UTF-8') as project_file:
    lines = project_file.readlines()
    for i in range(0, len(lines)):
        line = lines[i]
        if line == f"[ProjectVariant{assembly_count + 1}]\n":
            if lines[i+2].split('=')[0] == 'Description':
                assembly_count += 1
                assemblies.append(lines[i+2].split('=')[1].rstrip())

print('\nProject Information')
print(  '*******************')
print(str(assembly_count) + ' Assemblie(s) Found In Project File:')
for i in assemblies: print(i)
print()

###################################################################################################
####################     Declare arrays and dictionaries for needed files     #####################
###################################################################################################

# List of file we are going to be manipulating
###################################################################################################

# Assembly BOMs found
assembly_boms = []
# SAP BOMs found
sap_boms = []
# Aegis Sync BOMs found
aegis_boms = []

# Lists of files that we want to keep with known names
###################################################################################################

# List of files for the reports folder that we want to keep
reports_files = [pcb_number + ' Order Information.xlsx',
                 pcb_number + '_Build_Request.docx',
                 pcb_number + '_Build Request.docx',
                 pcb_number + ' Build_Request.docx',
                 pcb_number + ' Build Request.docx',
                 pcb_number + '_EE_Review.xlsx',
                 pcb_number + '_EE Review.xlsx',
                 pcb_number + '_EE Review.xlsx',
                 pcb_number + ' EE_Review.xlsx',
                 pcb_number + '_EE_Review.xls',
                 pcb_number + '_EE Review.xls',
                 pcb_number + '_EE Review.xls',
                 pcb_number + ' EE_Review.xls'
                 ]

# List of files for Source folder we want to keep, SCH HANDELED ON SEARCH
source_files = ['Assy_' + pcb_number + '.PCBDwf',
                'PCB_' + pcb_number + '.PcbDoc',
                ]

# List of Mfg-Data files to keep, adds aegis sync later depending on if excel or text needed.
mfgdata_files = ['ODB_' + pcb_number + '.zip'
                 ]

# Dictionary of Altiums Gerber layer file extensions
gerber_ext = {  'G' : '   -  Mid Layer ',              # .G(int) =  internal layer (int)
              'GBL' : '  -  Bottom Layer',
              'GBO' : '  -  Bottom Overlay',
              'GBP' : '  -  Bottom Paste Mask',
              'GBS' : '  -  Bottom Solder Mask',
#              'GD' : '   -  Drill Drawing ',        # .GD(int) = Drill Drawing (int)
#              'GG' : '   -  Drill Guide ',          # .GG(int) = Drill Guide (int)
#             'GKO' : '  -  Keep Out Layer',
#              'GM' : '  -  Mechanical Layer ',     # .GM(int) = Mechanical Layer (int)
               'GP' : '  -  Internal Plane Layer ', # .GP(int) = Internal Plane Layer (int)
#             'GPB' : '  -  Pad Master Bottom',
#             'GPT' : '  -  Pad Master Top',
              'GTL' : '  -  Top Layer',
              'GTO' : '  -  Top Overlay',
              'GTP' : '  -  Top Paste Mask',
              'GTS' : '  -  Top Solder Mask',
#             'P01' : '  -  Gerber Panels ',        # .P0(int) = Ger Panel (int)
              'DRR' : '  -  NC Drill Report',
              'TXT' : '  -  Drill File'
              }

# List of Gerber files extensions that repeat with numbers
rep_gerber_ext = ['G',
#                 'GD',
#                 'GG',
#                 'GM',
                  'GP',
#                 'P0'
                  ]

# List of Gerber files to keep, filled in on search. Includes the correct number suffex if appropriate
# Added hole and slot now to cover if present later
keep_gerbers = [pcb_number + '-SlotHoles.TXT',
                pcb_number + '-RoundHoles.TXT'
                ]

# List of files / folders for CAM folder that are valid
cam_files = ['Gerber and Drill',
              pcb_number + '.pdf',
              pcb_number + '.PDF',
              pcb_number + '.zip',
              'Spec_' + pcb_number + '.dwg',
             ]
             
###################################################################################################
######################     Find all the relavent project info and files     #######################
###################################################################################################

# Find the assembly BOMs
print('Assembly BOMs:')
# For each assembly, see if the its BOM is in the reports folder
for assembly in assemblies:
    if os.path.isfile('.\\Reports\\' + assembly + ' Assembly BOM.xlsx'):
        print(assembly + ' Assembly BOM Found')
        assembly_boms.append(assembly + ' Assembly BOM.xlsx')
        reports_files.append(assembly + ' Assembly BOM.xlsx')
    else:
        print(assembly + ' MISSING')
print()

# Find the SAP BOMs
print('SAP BOMs:')
# For each assembly, see if the its BOM is in the reports folder
for assembly in assemblies:
    if os.path.isfile('.\\Reports\\' + assembly + ' SAP Import File.xlsx'):
        print(assembly + ' SAP Import File Found')
        sap_boms.append(assembly + ' SAP Import File.xlsx')
        reports_files.append(assembly + ' SAP Import File.xlsx')
    else:
        print(assembly + ' MISSING')
print()

# Find Gerber layer data from Gerber and Drill folder
print('Gerber Files Found:')
# List Comprehension https://treyhunner.com/2015/12/python-list-comprehensions-now-in-color/
# For all files in the Gerber and Drill folder, if the extension is in the gerber extension dictionary, add it to the keep list.
for file in [file for file in os.listdir('..\\Cam\\Gerber and Drill\\') if os.path.isfile('..\\Cam\\Gerber and Drill\\' + file)]:
    #if the file name is the board number and the extension is in the dictionary
    if file.rsplit('.', 1)[0] == pcb_number and file.rsplit('.', 1)[1] in gerber_ext:
        print(file + gerber_ext[file.rsplit('.', 1)[1]])
        keep_gerbers.append(file)
    # Most of this checking if the extension starts with extension in list and ends with digit could be replaced with regular expression, also verifies the board number
    for extension in rep_gerber_ext:
        if file.rsplit('.', 1)[0] == pcb_number and file.rsplit('.', 1)[1].startswith(extension):
            # Remove the extension prefix to get number
            if file.rsplit('.', 1)[1].replace(extension, '').isdigit():
                print(file + gerber_ext[extension] + file.rsplit('.', 1)[1].replace(extension, ''))
                keep_gerbers.append(file)
print()

# Find the Aegis Sync BOMs
print('Aegis Sync Excel BOMs:')
# For each assembly, see if the its BOM is in the reports folder
for assembly in assemblies:
    # If both txt and excel files are there, see which one is newer
    if os.path.isfile('..\\Mfg-Data\\Aegis Sync_' + assembly + '.xlsx') and os.path.isfile('..\\Mfg-Data\\Aegis Sync_' + assembly + '.txt'):
        # If the excel is newer, alert user and save only excel to be saved
        if os.path.getmtime('..\\Mfg-Data\\Aegis Sync_' + assembly + '.xlsx') >= os.path.getmtime('..\\Mfg-Data\\Aegis Sync_' + assembly + '.txt'):
            print(assembly + '.txt Aegis Sync File Found - OUT OF DATE')
            print(assembly + '.xlsx Aegis Sync File Found')
            aegis_boms.append('Aegis Sync_' + assembly + '.xlsx')
            mfgdata_files.append('Aegis Sync_' + assembly + '.xlsx')
        # If the text is newer, alert user and save only excel to be saved
        else:
            print(assembly + '.txt Aegis Sync File Found')
            print(assembly + '.xlsx Aegis Sync File Found - Marked For Deletion')
            mfgdata_files.append('Aegis Sync_' + assembly + '.txt')
    # If just the excel file is there add it to be saved
    elif os.path.isfile('..\\Mfg-Data\\Aegis Sync_' + assembly + '.xlsx'):
        print(assembly + '.xlsx Aegis Sync File Found')
        aegis_boms.append('Aegis Sync_' + assembly + '.xlsx')
        mfgdata_files.append('Aegis Sync_' + assembly + '.xlsx')
    # Just the text BOM was found
    elif os.path.isfile('..\\Mfg-Data\\Aegis Sync_' + assembly + '.txt'):
        print(assembly + '.txt Aegis Sync File Found')
        mfgdata_files.append('Aegis Sync_' + assembly + '.txt')
    else:
        print(assembly + ' MISSING')
    
###################################################################################################
#################################     Sort the assembly BOMs     ##################################
###################################################################################################

print('\nSorting Assembly BOMs')
print('*********************')

# Ask the user if they would like to sort all assembly BOM
response_all = ''
if len(assembly_boms) > 1:
    while response_all not in ('y', 'yes', 'n', 'no'):
        response_all = input('Would you like to sort all assembly BOMs?: ')
        print()
else:
    response_all = 'no'

# Do it for all assembly BOMs found
for assembly_bom in assembly_boms:

    # Ask the user if they would like to sort each assembly BOM if they didnt want to do all
    response2 = ''
    if re.search('^n(o)?$', response_all, re.IGNORECASE):
            while response2 not in ('y', 'yes', 'n', 'no'):
                response2 = input(f'Would you like to sort {assembly_bom}?: ')
    if re.search('^n(o)?$', response2, re.IGNORECASE) and re.search('^n(o)?$', response_all, re.IGNORECASE):
        print(assembly_bom + ' Skipped\n')
        continue
    
    print('Sorting ' + assembly_bom + '...')
    # Open the excel file and go in the first sheet
    wb = openpyxl.load_workbook('.\\Reports\\' + assembly_bom)
    sheet = wb['Sheet1']

    # Set up fonts
    # Font for component section declaration 
    comp_row_font  = NamedStyle(name = 'comp_row_font')
    comp_row_font.font = Font(name = 'Verdana', bold = True, size = 14)
    # Font for the section headers
    comp_header_font  = NamedStyle(name = 'comp_header_font')
    comp_header_font.font = Font(name = 'Verdana', bold = True, size = 12, underline = 'single')
    # Font for default lines
    default_font  = NamedStyle(name = 'default_font')
    default_font.font = Font(name = 'Verdana', size = 12)
    # Something to tell when the BOM starts
    header_row = 0
    # A list that will contain dictionaries of the parts
    bom_content = []

# Read the cells
###################################################################

    # Scan through each row in the BOM excel from 1 to end
    for row in range(1, sheet.max_row + 1):
        # Column 1 is part number, 2 is quantity etc. Store the values in these variables
        part_number = sheet.cell(row, 1).value
        quantity    = sheet.cell(row, 2).value
        description = sheet.cell(row, 3).value
        designator  = sheet.cell(row, 4).value
        layer       = sheet.cell(row, 5).value
        fitted      = sheet.cell(row, 6).value
        # If its previously found the header, and the part number is not None, add the part
        if header_row != 0 and part_number is not None:
            # Handle layer None components, ask for input until valid response
            if layer == 'None':
                while 1:
                    response = input(f'{part_number} | {description} | is layer None, does it belong on top or bottom side?: ')
                    if re.search('^t(op)?( )?(side)?$', response, re.IGNORECASE):
                        layer = 'Top'
                        break
                    if re.search('^b(ottom)?( )?(side)?$', response, re.IGNORECASE):
                         layer = 'Bottom'
                         break
            # Add the part as a dictionary to the list of parts. Make them strings to allow sorting to work.
            bom_content.append({'part_number': str(part_number),
                                'quantity': quantity,
                                'description': str(description),
                                'designator': str(designator),
                                'layer': str(layer),
                                'fitted': str(fitted)})
        # If the row equals the header Altium uses, singal that the header has been found
        if str(part_number) + str(quantity) + str(description) + str(designator) + str(layer) + str(fitted) == 'LibRefQuantityDescriptionDesignatorLayerFitted':
            header_row = row
    # If the header was not found, it was probably already messed with. Alert the user
    if header_row == 0:
        print(f'No Altium header row found for {assembly_bom}. Unable to sort...\n')
        continue
    
    # Sort the BOMs
    bom_content.sort(key = lambda k: k['part_number'])
    bom_content.sort(key = lambda k: k['fitted'])
    bom_content.sort(key = lambda k: k['layer'], reverse = True)

    # Delete the old rows of data 
    sheet.delete_rows(header_row, sheet.max_row)
    
# Write the data to the cells
###################################################################

    current_section = 'None'
    pattern_fill = False
    # Start printing 14 rows up from where the header row was found
    row = header_row - 14
    for part in bom_content:
        # Check if you are in a new section of the BOM, if so print the header
        if current_section != part['layer'] + part['fitted']:
            row += 2
            current_section = part['layer'] + part['fitted']
            # Determine what section of BOM is starting
            if part['layer'] == 'Top': header_string = 'TOP SIDE COMPONENTS'
            elif part['layer'] == 'Bottom': header_string = 'BOTTOM SIDE COMPONENTS'
            else: header_string = 'None Side Components'
            if part['fitted'] == 'Not Fitted': header_string += ' (not installed)'
            # Print the component section row and format it
            sheet.cell(row = row, column = 1).value = header_string
            sheet['A' + str(row)].style = comp_row_font
            # Print the component headers and format it
            sheet.cell(row = row + 1, column = 1).value = 'PART#'
            sheet['A' + str(row + 1)].style = comp_header_font
            sheet.cell(row = row + 1, column = 2).value = 'QTY'
            sheet['B' + str(row + 1)].style = comp_header_font
            sheet.cell(row = row + 1, column = 3).value = 'DESCRIPTION'
            sheet['C' + str(row + 1)].style = comp_header_font
            sheet.cell(row = row + 1, column = 4).value = 'SYMBOL'
            sheet['D' + str(row + 1)].style = comp_header_font
            row += 2
            pattern_fill = False
        # Print the row data for the component, give them default 
        sheet.cell(row = row, column = 1).value = part['part_number']
        sheet['A' + str(row)].style = default_font
        if pattern_fill == True: sheet['A' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 2).value = part['quantity']
        sheet['B' + str(row)].style = default_font
        if pattern_fill == True: sheet['B' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 3).value = part['description']
        sheet['C' + str(row)].style = default_font
        if pattern_fill == True: sheet['C' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 4).value = part['designator']  
        sheet['D' + str(row)].style = default_font
        if pattern_fill == True: sheet['D' + str(row)].fill = PatternFill('solid', fgColor='D9D9D9')
        sheet.cell(row = row, column = 5).value = part['layer']
        sheet['E' + str(row)].style = default_font
        sheet.cell(row = row, column = 6).value = part['fitted']
        sheet['F' + str(row)].style = default_font
        row += 1
        pattern_fill = not pattern_fill

    # Save the excel
    wb.save('.\\Reports\\' + assembly_bom)
    print(assembly_bom + ' sorting complete\n')

###################################################################################################
#################################     Clean SAP Import Files     ##################################
###################################################################################################

print('Cleaning SAP Import Files')
print('*************************')

# Ask the user if they would like to sort all assembly BOM
response = ''
if len(sap_boms) > 1:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('Would you like to clean all SAP Import files?: ')
        print()
else:
    response = 'no'

# Do it for all assembly BOMs found
for sap_bom in sap_boms:

    # Ask the user if they would like to sort each assembly BOM if they didnt want to do all
    response2 = ''
    if re.search('^n(o)?$', response, re.IGNORECASE):
            while response2 not in ('y', 'yes', 'n', 'no'):
                response2 = input(f'Would you like to clean {sap_bom}?: ')
    if re.search('^n(o)?$', response2, re.IGNORECASE) and re.search('^n(o)?$', response, re.IGNORECASE):
        print(sap_bom + ' Skipped\n')
        continue
    
    print('Cleaning ' + sap_bom + '...')
    # Open the excel file and go in the first sheet
    wb = openpyxl.load_workbook('.\\Reports\\' + sap_bom)
    sheet = wb['Sheet1']

    # Verify the header row is there
    if sheet.cell(1, 2).value != 'LibRef' or sheet.cell(1, 4).value != 'Quantity':
        print(f'No Altium header row found for {sap_bom}. Unable to sort...\n')
        continue

    # Move all cells up to delete header
    for row in range(1, sheet.max_row + 1):
        for column in range(1, 5):
            sheet.cell(row = row, column = column).value = sheet.cell(row = row + 1, column = column).value 

    # Save the data
    wb.save('.\\Reports\\' + sap_bom)
    print(sap_bom + ' cleaning complete\n')

###################################################################################################
#################################     Tab Delimit Aegis Sync     ##################################
###################################################################################################

print('Tab Delimit Aegis Sync Files')
print('****************************') 

# Ask the user if they would like to sort all assembly BOM
response = ''
if len(aegis_boms) > 1:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('Would you like to create all Aegis sync files?: ')
        print()
else:
    response = 'no'

# Do it for all assembly BOMs found
for excel in aegis_boms:

    # Ask the user if they would like to sort each assembly BOM if they didn't want to do all
    response2 = ''
    if re.search('^n(o)?$', response, re.IGNORECASE):
            while response2 not in ('y', 'yes', 'n', 'no'):
                response2 = input(f'Would you like to make text file for {excel}?: ')
    if re.search('^n(o)?$', response2, re.IGNORECASE) and re.search('^n(o)?$', response, re.IGNORECASE):
        print(sap_bom + ' Skipped\n')
        continue

    # If the text file already exists, delete it
    if os.path.isfile('..\\Mfg-Data\\' + excel.rsplit('.', 1)[0] + '.txt'):
        try:
            os.remove('..\\Mfg-Data\\' + excel.rsplit('.', 1)[0] + '.txt')
        except OSError as e: 
            print ("Error: %s - %s." % (e.filename, e.strerror))
            
    # Open the excel file and go in the first sheet
    print('Generating text file for ' + excel)
    wb = openpyxl.load_workbook('..\\Mfg-Data\\' + excel)
    sheet = wb['Sheet1']
    # Save each row to text file, tab delimited
    with open('..\\Mfg-Data\\' + excel.rsplit('.', 1)[0] + '.txt', 'w', newline = '') as aegis_text:
        file = csv.writer(aegis_text, delimiter = '\t')
        for row in sheet.rows:
            file.writerow([cell.value for cell in row])

    # Finished with conversion, no need to save excel
    print(excel + ' converted to text\n')

    # If the excel is in mfgdata_files, replace it with the text
    mfgdata_files = [excel.rsplit('.', 1)[0] + '.txt' if x == (excel) else x for x in mfgdata_files]
    
###################################################################################################
###################################     Remove Junk Files     #####################################
###################################################################################################

print('Deleting Unneeded Files')
print('***********************\n')

# Reports Folder
###################################################################################################

print('Reports Folder')
print('**************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
reports_unneeded = []

# Prints all the files in Reports folder that are not in the list to keep
for unneeded_file in [file for file in os.listdir('.\\Reports\\') if file not in reports_files]:
    reports_unneeded.append(unneeded_file)
    print(unneeded_file)

# Ask if ok to delete all unneeded file from Reports
response = ''
if reports_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Reports?')
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in reports_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('.\\Reports\\' + unneeded_file):
                try:
                    os.remove('.\\Reports\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    shutil.rmtree('.\\Reports\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Reports folder done\n')

# Source Folder
###################################################################################################

print('Source Folder')
print('*************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
source_unneeded = []

# Danger, mess ahead. RE searches with variable names would fix a lot
# Finds all unneeded files that aren't in the list, schematic docs, and handles _vO1 files
for file in os.listdir('.\\Source\\'):
    # Dont even continue if its already a valid file
    if file not in source_files:
        # if the item is a file, keep checking, if its a folder delete
        if os.path.isfile('.\\Source\\' + file):
            # the file is schematic dont add it
            if not re.search('schdoc', file.rsplit('.', 1)[1], re.IGNORECASE):
                # for non schematic files, if has _v01 in the name keep checking, otherwise add it
                if re.search('_v[0-9]+$', file.rsplit('.', 1)[0], re.IGNORECASE):
                    # if removing _v01 still doesnt match, add it
                    if file.rsplit('_', 1)[0] + '.' + file.rsplit('.', 1)[1] not in source_files:
                        source_unneeded.append(file)
                else:
                    source_unneeded.append(file)
        else:
            source_unneeded.append(file)
# Print all files 
for unneeded_file in source_unneeded:
    print(unneeded_file)

# Ask if OK to delete all unneeded file from Reports
response = ''
if source_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Source?')
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in source_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('.\\Source\\' + unneeded_file):
                try:
                    os.remove('.\\Source\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    shutil.rmtree('.\\Source\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Source folder done\n')

# CAM Folder
###################################################################################################

print('CAM Folder')
print('**************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
cam_unneeded = []

# Danger, mess ahead. RE searches with variable names would fix a lot
# Finds all unneeded files that aren't in the list, schematic docs, and handles _vO1 files
for file in os.listdir('..\\Cam\\'):
    # Don t even continue if its already a valid file, handles valid folder names here
    if file not in cam_files:
        # If the item is a file, keep checking, if its a folder delete
        if os.path.isfile('..\\Cam\\' + file):
            # If the file starts with a board number
            if re.search('^' + pcb_number, file, re.IGNORECASE):
                # And not .zip or pdf, delete it
                if not (re.search('.zip$', file, re.IGNORECASE) or re.search('.pdf$', file, re.IGNORECASE)):
                    cam_unneeded.append(file)
            # Doesnt start with a board number but need to start with spec and be a drawing
            elif not (re.search('^Spec_', file, re.IGNORECASE) or re.search('.dwg$', file, re.IGNORECASE)):
                cam_unneeded.append(file)
        else:
            cam_unneeded.append(file)

# Print all files 
for unneeded_file in cam_unneeded:
    print(unneeded_file)

# Ask if OK to delete all unneeded file from Reports
response = ''
if cam_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Reports?')
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in cam_unneeded:
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('..\\Cam\\' + unneeded_file):
                try:
                    os.remove('..\\Cam\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    shutil.rmtree('..\\Cam\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Cam folder done\n')

# Gerber and Drill Folder
###################################################################################################

print('Gerber and Drill Folder')
print('***********************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
gerber_unneeded = []

# Prints all the files in Reports folder that are not in the list to keep
for unneeded_file in [file for file in os.listdir('..\\Cam\\Gerber and Drill') if file not in keep_gerbers]:
    gerber_unneeded.append(unneeded_file)
    print(unneeded_file)

# Ask if OK to delete all unneeded file from Reports
response = ''
if gerber_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Gerber and Drill?')
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in gerber_unneeded:
            print(unneeded_file)
            # Verify its a file and if so try to delete, if not report back
            if os.path.isfile('..\\Cam\\Gerber and Drill\\' + unneeded_file):
                try:
                    os.remove('..\\Cam\\Gerber and Drill\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    shutil.rmtree('..\\Cam\\Gerber and Drill\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Gerber and Drill folder done\n')

# Manufacturing and Data Folder
###################################################################################################

print('Mfg-Data Folder')
print('***************')

# Array to store files that it finds unneeded. Ensures whats deleted is what user saw
mfg_unneeded = []

# Prints all the files in Reports folder that are not in the list to keep
for unneeded_file in [file for file in os.listdir('..\\Mfg-Data\\') if file not in mfgdata_files]:
    mfg_unneeded.append(unneeded_file)
    print(unneeded_file)

# Ask if ok to delete all unneeded file from Reports
response = ''
if mfg_unneeded:
    while response not in ('y', 'yes', 'n', 'no'):
        response = input('OK to delete all unneeded files from Mfg-Data?')
    if re.search('^y(es)?$', response, re.IGNORECASE):
        # Delete all the files
        for unneeded_file in mfg_unneeded:
            # Verify its a file and if so try to delete, if not report back
            print('..\\Mfg-Data\\' + unneeded_file)
            if os.path.isfile('..\\Mfg-Data\\' + unneeded_file):
                try:
                    os.remove('..\\Mfg-Data\\' + unneeded_file)
                except OSError as e: 
                    print ("Error: %s - %s." % (e.filename, e.strerror))
            else:
                try:
                    shutil.rmtree('..\\Mfg-Data\\' + unneeded_file)
                except OSError as e:
                    print ("Error: %s - %s." % (e.filename, e.strerror))
print('Mfg-Data folder done\n')

# Change working directory back to default to prevent program from preventing deleting project file
os.chdir(owd)
print('Documentation Cleanup Complete!!')
